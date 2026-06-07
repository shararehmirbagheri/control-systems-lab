from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font

EAM_FOLDER = "EAM_UDT_Definitions_20260605"
IAD_FOLDER = "IAD2_DC2_UDT_Definitions_20260605"

OUTPUT_FILE = "folder_comparison.xlsx"


def get_items(folder):
    return sorted(
        str(p.relative_to(folder))
        for p in folder.rglob("*")
    )


base = Path(__file__).parent

eam_path = base / EAM_FOLDER
iad_path = base / IAD_FOLDER

eam_items = get_items(eam_path)
iad_items = get_items(iad_path)

wb = Workbook()
ws = wb.active
ws.title = "Comparison"

ws.append(["EAM Item", "IAD2 Item", "Status"])

red_font = Font(color="FF0000")

all_items = sorted(set(eam_items) | set(iad_items))

for item in all_items:

    in_eam = item in eam_items
    in_iad = item in iad_items

    if in_eam and in_iad:
        ws.append([item, item, "Match"])

    elif in_eam:
        row = ws.max_row + 1
        ws.append([item, "Does Not Exist In IAD2", "Missing in IAD2"])

        ws.cell(row=row, column=2).font = red_font
        ws.cell(row=row, column=3).font = red_font

    elif in_iad:
        row = ws.max_row + 1
        ws.append(["Does Not Exist In EAM", item, "Missing in EAM"])

        ws.cell(row=row, column=1).font = red_font
        ws.cell(row=row, column=3).font = red_font

for column in ws.columns:
    max_length = max(len(str(cell.value)) if cell.value else 0 for cell in column)
    ws.column_dimensions[column[0].column_letter].width = max_length + 5

wb.save(base / OUTPUT_FILE)

print(f"Excel file created: {OUTPUT_FILE}")